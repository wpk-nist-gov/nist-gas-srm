from __future__ import annotations

import re
from contextlib import contextmanager
from functools import partial
from typing import TYPE_CHECKING, cast

import pandas as pd
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import (
    dataframe_to_rows,  # pyright: ignore[reportUnknownVariableType]
)

if TYPE_CHECKING:
    from collections.abc import Callable, Container, Generator, Sequence
    from io import BytesIO
    from pathlib import Path
    from typing import Any

    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.styles.fills import PatternFill
    from openpyxl.worksheet.worksheet import Worksheet


EXCEL_FILENAME_PATTERN = re.compile(
    r"srm(?P<srm_id>\d+)(?P<batch_id>\w*)_Series(?P<lot_id>\w*)_(.*).xls",
    flags=re.IGNORECASE,
)

strip_trailing_numbers = partial(re.compile(r"\.[1-9]+").sub, "")


def skipper(
    lower: int | None = None,
    upper: int | None = None,
    include: Container[int] | None = None,
) -> Callable[[int], bool]:
    def func(x: int) -> bool:
        return (
            (lower is not None and x < lower)
            or (upper is not None and x > upper)
            or (include is not None and x not in include)
        )

    return func


def maybe_dropna(df: pd.DataFrame | None, **kwargs: Any) -> pd.DataFrame | None:
    if df is not None:
        return cast("pd.DataFrame", df.dropna(**kwargs))
    return df


def get_frame(
    io: Any,
    sheet_name: str,
    **kwargs: Any,
) -> pd.DataFrame:
    return pd.read_excel(io, sheet_name=sheet_name, **kwargs).dropna(how="all")  # pyright: ignore[reportUnknownMemberType]


def get_value_from_worksheet(
    xls: pd.ExcelFile, sheet_name: str, rowx: int, colx: int | str
) -> Any:
    if isinstance(colx, str):
        colx = column_index_from_string(colx) - 1

    book: Any = xls.book  # pyright: ignore[reportUnknownVariableType, reportUnknownMemberType]
    engine = cast("str | None", getattr(xls, "engine", None))
    if engine == "xlrd":
        return book.sheet_by_name(sheet_name).cell_value(rowx=rowx, colx=colx)  # pyright: ignore[reportUnknownMemberType, reportUnknownVariableType]

    if engine == "openpyxl":
        return cast(
            "int",
            book  # pyright: ignore[reportUnknownMemberType]
            .get_sheet_by_name(sheet_name)
            .cell(row=rowx + 1, column=colx + 1)
            .value,
        )

    if engine == "calamine":
        return book.get_sheet_by_name(sheet_name).to_python(skip_empty_area=False)[  # pyright: ignore[reportUnknownMemberType, reportUnknownVariableType]
            rowx
        ][colx]

    msg = f"Unknown engine={engine}"
    raise ValueError(msg)


@contextmanager
def as_excelfile(
    path_or_excelfile: Path | BytesIO | pd.ExcelFile,
) -> Generator[pd.ExcelFile]:
    if isinstance(path_or_excelfile, pd.ExcelFile):
        yield path_or_excelfile
    else:
        yield pd.ExcelFile(path_or_excelfile)


def get_frame_with_len_check(
    path_or_excelfile: Path | BytesIO | pd.ExcelFile,
    sheet_name: str,
    usecols: str,
    rowx: int,
    colx: int | str,
    require_check: bool = True,
    **kwargs: Any,
) -> pd.DataFrame:

    with as_excelfile(path_or_excelfile) as xls:
        df = get_frame(xls, sheet_name=sheet_name, usecols=usecols, **kwargs)

        if require_check:
            if (
                val := get_value_from_worksheet(
                    xls, sheet_name=sheet_name, rowx=rowx, colx=colx
                )
            ) is None:
                msg = "No check value found"
                raise ValueError(msg)

            if (check := int(val)) != len(df):
                msg = f"Wrong check shape {check=} != {len(df)}"
                raise ValueError(msg)

    return df


def parse_excel_filename_to_metadata(name: str) -> dict[str, Any]:
    """
    Parse an excel filename to parameters
    """
    if (m := EXCEL_FILENAME_PATTERN.match(name)) is None:
        msg = f"Unable to parse ids from {name}"
        raise ValueError(msg)

    out = m.groupdict().copy()
    if not out["batch_id"]:
        out["batch_id"] = None

    return out


def optional_dataframe_func_wrapper(
    func: Callable[..., pd.DataFrame],
    xls: pd.ExcelFile,
    sheet_name: str,
    **kwargs: Any,
) -> pd.DataFrame | None:

    if sheet_name in xls.sheet_names:
        df = func(
            xls,
            sheet_name,
            **kwargs,
        )
        return None if df.empty else df
    return None


# * openpyxl
def get_fill_from_cell(cell: Cell | MergedCell) -> PatternFill:
    from copy import copy

    return cast("PatternFill", copy(cell.fill))


def validate_column(col: int | str) -> int:
    if isinstance(col, str):
        return column_index_from_string(col)
    return col


def _validate_start(start: tuple[int, int | str]) -> tuple[int, int]:
    row, col = start
    if isinstance(col, str):
        col = column_index_from_string(col)
    return (row, validate_column(col))


def simple_write_to_excel(
    obj: pd.DataFrame,
    worksheet: Worksheet,
    index: bool = False,
    header: bool = True,
    start: tuple[int, int | str] = (1, 1),
    fill_from: tuple[int, int | str] | None = (2, 1),
    rows: Sequence[int] | None = None,
    columns: Sequence[int | str] | None = None,
) -> None:
    if obj.empty:
        return

    start = _validate_start(start)
    fill = (
        get_fill_from_cell(worksheet.cell(*_validate_start(fill_from)))
        if fill_from is not None
        else None
    )

    row_start, col_start = start

    columns_strict: Sequence[int] = (
        range(col_start, obj.shape[1] + col_start)
        if columns is None
        else [validate_column(col) for col in columns]
    )
    if rows is None:
        rows = range(row_start, obj.shape[0] + row_start + int(header))

    for r_idx, row in zip(
        rows, dataframe_to_rows(obj, index=index, header=header), strict=True
    ):
        for c_idx, value in zip(columns_strict, row, strict=True):
            target_cell = cast("Cell", worksheet.cell(row=r_idx, column=c_idx))
            target_cell.value = value
            if fill is not None:
                target_cell.fill = fill
